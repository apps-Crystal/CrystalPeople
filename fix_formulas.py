from openpyxl import load_workbook

wb = load_workbook('/sessions/loving-adoring-clarke/mnt/PMS/Crystal_People_SOP_Tracker.xlsx')
ws = wb['SOP Build Status']

# Fix phase status formulas (rows 98-106)
for row in range(98, 107):
    c = ws.cell(row=row, column=5)
    r = row
    c.value = f'=IF(C{r}=0,"N/A",IF(C{r}=D{r},"Complete",IF(D{r}=0,"Not Started","In Progress")))'

wb.save('/sessions/loving-adoring-clarke/mnt/PMS/Crystal_People_SOP_Tracker.xlsx')
print('Fixed')
