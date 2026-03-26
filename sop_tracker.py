from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "SOP Build Status"

# Colors
HEADER_FILL = PatternFill('solid', fgColor='1B3A5C')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
SECTION_FONT = Font(name='Arial', bold=True, size=10, color='1B3A5C')
DONE_FILL = PatternFill('solid', fgColor='C6EFCE')
PARTIAL_FILL = PatternFill('solid', fgColor='FFEB9C')
NOT_FILL = PatternFill('solid', fgColor='FFC7CE')
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
thin = Side(style='thin', color='B0B0B0')
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

# Column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 52
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 45

# Title row
ws.merge_cells('A1:F1')
ws['A1'] = 'Crystal People — SOP vs Build Status Tracker'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1B3A5C')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:F2')
ws['A2'] = 'SOP Document: HR-SOP-002 v1.0  |  As of: 26 March 2026'
ws['A2'].font = Font(name='Arial', size=9, italic=True, color='666666')
ws['A2'].alignment = Alignment(horizontal='center')

# Headers
headers = ['S.No', 'SOP Section', 'Feature / Requirement', 'Status', 'Phase', 'Remarks']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
ws.row_dimensions[4].height = 24

data = [
    # Section 3
    ('S', '3. Roles', 'User Roles & Permissions', '', '', ''),
    (1, '3', '4 roles: Employee, Manager, HR, MD/CEO', 'Done', '1', 'JWT auth with role-based access'),
    (2, '3', 'Role-based visibility (Can See / Can Do per role table)', 'Done', '1', 'Enforced in all API routes + sidebar'),
    (3, '3', 'Key control: Employee cannot self-approve goals/ratings', 'Done', '2C', 'Manager and employee always different'),

    # Section 4
    ('S', '4. Employee Types', 'Two Employee Types', '', '', ''),
    (4, '4.1', 'White collar — scored against monthly goals', 'Done', '2C', 'Goals + 3 behaviours'),
    (5, '4.1', 'Min 3, max 5 goals at all times', 'Done', '2C', 'Validated in goal API'),
    (6, '4.1', 'Cannot drop below 3 without replacement', 'Done', '2C', 'Enforced in drop goal logic'),
    (7, '4.1', 'Dropped goals require mandatory written reason', 'Done', '2C', 'Drop_Reason field required'),
    (8, '4.1', 'New mid-month goals active immediately, scored next month', 'Done', '2C', 'Created_At tracked'),
    (9, '4.2', 'Blue collar — fixed task scorecard', 'Done', '2C', 'Tasks sheet + scoring'),
    (10, '4.2', '3–5 tasks per employee, fixed for year', 'Done', '2C', 'Manager sets via Team Tasks UI'),
    (11, '4.2', 'Each task scored 1–5 monthly by manager', 'Done', '2C', 'Monthly_Scores JSON in Tasks sheet'),
    (12, '4.3', '3 behavioural dimensions (Attendance, Teamwork, Ownership)', 'Done', '2C', 'BEHAVIOUR_DIMENSIONS constant'),

    # Section 5
    ('S', '5. Rating Scale', 'The Rating Scale', '', '', ''),
    (13, '5', '1–5 scale with labels (Outstanding → Unsatisfactory)', 'Done', '2C', 'getRatingBand() in utils.ts'),
    (14, '5', 'Simple average of all dimensions, no weighting', 'Done', '2C', 'computeAverage() in utils.ts'),

    # Section 6
    ('S', '6. Weekly Cycle', 'The Weekly Cycle — Every Friday', '', '', ''),
    (15, '6.1', 'Employee weekly reflection (4 fields + mood)', 'Done', '2B', 'Weekly_Reflections sheet'),
    (16, '6.2', 'Manager weekly check-in (5 fields with visibility rules)', 'Done', '2B', 'Weekly_Checkins sheet'),
    (17, '6.2', 'Concern field visible to HR/MD only, never to employee', 'Done', '2B', 'Filtered in API response'),
    (18, '6.3', 'Employee acknowledgement of check-in', 'Done', '2B', 'Acknowledged_At in reflections'),
    (19, '6.3', 'Saturday auto reminder to non-submitters', 'Not Built', '3', 'Needs scheduled job'),
    (20, '6.3', 'Monday — pending visible to HR', 'Not Built', '3', 'Needs escalation dashboard'),
    (21, '6.3', 'Missed reflections tracked, visible to MD', 'Not Built', '3', 'Needs exec escalation view'),

    # Section 7
    ('S', '7. Monthly Cycle', 'The Monthly Cycle', '', '', ''),
    (22, '7', 'Day 1 — Review window opens, employees & managers notified', 'Partial', '3', 'Window logic exists, auto-notify not built'),
    (23, '7', 'Day 1–5 — Employee self-scores all dimensions', 'Done', '2C', 'Self-score page + API'),
    (24, '7', 'Day 1–10 — Manager scores each team member', 'Done', '2C', 'Score-team page + API'),
    (25, '7', 'After scoring — Manager writes review notes + next month focus', 'Done', '2D', 'Stored in Review_Cycles'),
    (26, '7', 'After manager submits — Employee acknowledges review', 'Done', '2D', 'Acknowledge API + My Review page'),
    (27, '7', 'Day 3 — First auto reminder to non-scorers', 'Not Built', '3', 'Needs scheduled job'),
    (28, '7', 'Day 6 — Second auto reminder', 'Not Built', '3', 'Needs scheduled job'),
    (29, '7', 'Day 7 — Pending reviews visible to HR', 'Partial', '2D', 'HR Monitor page shows all statuses'),
    (30, '7', 'Day 10 — Pending visible to MD, affects manager\'s own score', 'Not Built', '3', 'Needs exec dashboard + penalty logic'),
    (31, '7', 'Day 11 — Window closes, scores locked for the month', 'Not Built', '3', 'Manual lock exists, auto-lock not built'),
    (32, '7.1', 'Self-score: ≥2 point gap flagged for discussion', 'Done', '2D', 'Flagged_Dimensions in Review_Cycles'),
    (33, '7.2', 'AI validation of review notes (harsh/vague/unconstructive)', 'Done', '2D', 'ai-review.ts + validate-notes API'),
    (34, '7.3', 'Weekly context shown in monthly scoring screen', 'Done', '2D', 'weekly-context API + collapsible UI'),

    # Section 8
    ('S', '8. Increments', 'Annual Increment Process', '', '', ''),
    (35, '8', '12-month average score calculation', 'Not Built', '5', ''),
    (36, '8', 'MD sets increment percentage bands before April 1', 'Not Built', '5', ''),
    (37, '8', 'HR computes individual increments per band', 'Not Built', '5', ''),
    (38, '8', 'MD approves full increment sheet', 'Not Built', '5', ''),
    (39, '8', 'Increment letter generation', 'Not Built', '5', ''),

    # Section 9
    ('S', '9. First Login', 'First Login Experience', '', '', ''),
    (40, '9', 'One-time welcome screen explaining Crystal People', 'Done', '1', 'First_Login_Done flag + password change'),

    # Section 10
    ('S', '10. Org Mgmt', 'Org Management', '', '', ''),
    (41, '10', 'Excel upload for bulk employee setup/changes', 'Not Built', '4', ''),
    (42, '10', 'Quick edit — HR searches and edits any employee field', 'Done', '2A', 'Employee edit page + PATCH API'),
    (43, '10', 'New joiner before 15th — current month cycle', 'Not Built', '6', 'Edge case logic'),
    (44, '10', 'New joiner after 15th — starts next month', 'Not Built', '6', 'Edge case logic'),
    (45, '10', 'Employee exit — marked inactive, history preserved', 'Partial', '2A', 'Status field exists, no dedicated exit flow'),
    (46, '10', 'Reporting change — effective next month', 'Not Built', '6', 'Edge case logic'),
    (47, '10', 'Manager exits — team flagged unassigned', 'Not Built', '6', 'Edge case logic'),
    (48, '10', 'All org changes auto-logged (what, who, when)', 'Done', '2A', 'Org_Changes sheet + API'),

    # Section 11
    ('S', '11. Controls', 'Key Controls', '', '', ''),
    (49, '11', 'No employee may score themselves without manager scoring', 'Done', '2C', 'Status flow enforced'),
    (50, '11', 'No increment without MD written approval', 'Not Built', '5', 'Increment feature pending'),
    (51, '11', 'All cycles must be completed + acknowledged before lock', 'Done', '2D', 'Lock requires manager_scored or acknowledged'),
    (52, '11', 'Manager penalty if team reviews not done by Day 10', 'Not Built', '3', 'Needs penalty logic'),
    (53, '11', 'Rating 1 employees = no increment (MD exception only)', 'Not Built', '5', 'Increment feature pending'),
    (54, '11', 'Increment amounts strictly confidential between employees', 'Not Built', '5', 'Increment feature pending'),
    (55, '11', 'Manager concern flags never visible to employee', 'Done', '2B', 'Filtered in API'),

    # Section 12
    ('S', '12. Escalation', 'Escalation Matrix', '', '', ''),
    (56, '12', 'Employee disputes rating — HR mediates within 3 days', 'Not Built', '6', ''),
    (57, '12', 'Manager misses review deadline — HR escalates to dept head', 'Not Built', '3', ''),
    (58, '12', 'Employee misses reflection — auto reminder + HR visibility', 'Not Built', '3', ''),
    (59, '12', 'Goal/task change request — HR facilitates discussion', 'Not Built', '6', ''),
    (60, '12', 'Increment budget exceeded — HR presents to MD', 'Not Built', '5', ''),
    (61, '12', 'Employee refuses to acknowledge — HR documents, MD notified', 'Not Built', '6', ''),

    # Section 13
    ('S', '13. Backend', 'Google Sheets Backend Structure', '', '', ''),
    (62, '13', 'Employees sheet', 'Done', '1', '14 columns'),
    (63, '13', 'Managers sheet', 'Done', '1', 'Manager-Employee lookup'),
    (64, '13', 'Review_Cycles sheet', 'Done', '2C', '15 columns incl. AI flags'),
    (65, '13', 'Goals sheet', 'Done', '2C', '13 columns'),
    (66, '13', 'Tasks sheet', 'Done', '2C', '9 columns'),
    (67, '13', 'Weekly_Reflections sheet', 'Done', '2B', '9 columns'),
    (68, '13', 'Weekly_Checkins sheet', 'Done', '2B', '10 columns'),
    (69, '13', 'Org_Changes sheet', 'Done', '2A', '7 columns'),
    (70, '13', 'Notifications sheet', 'Done', '2D', '7 columns'),
    (71, '13', 'Config sheet', 'Done', '1', 'Key-Value pairs'),
    (72, '13', 'SEQUENCES sheet', 'Done', '1', 'Auto-increment IDs'),
]

row = 5
sno = 0
for item in data:
    if item[0] == 'S':
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=f"SOP {item[1]} — {item[2]}")
        c.font = SECTION_FONT
        c.fill = SECTION_FILL
        c.alignment = Alignment(vertical='center')
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
            ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.row_dimensions[row].height = 22
    else:
        sno += 1
        vals = [item[0], item[1], item[2], item[3], item[4], item[5]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if col == 1:
                c.alignment = Alignment(horizontal='center', vertical='center')
            if col == 4:
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.font = Font(name='Arial', size=10, bold=True)
                if v == 'Done':
                    c.fill = DONE_FILL
                    c.font = Font(name='Arial', size=10, bold=True, color='006100')
                elif v == 'Partial':
                    c.fill = PARTIAL_FILL
                    c.font = Font(name='Arial', size=10, bold=True, color='9C6500')
                elif v == 'Not Built':
                    c.fill = NOT_FILL
                    c.font = Font(name='Arial', size=10, bold=True, color='9C0006')
            if col == 5:
                c.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

# Summary section
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
c = ws.cell(row=row, column=1, value='Summary')
c.font = Font(name='Arial', bold=True, size=12, color='1B3A5C')
row += 1

summary = [
    ('Done', '=COUNTIF(D5:D200,"Done")', DONE_FILL, '006100'),
    ('Partial', '=COUNTIF(D5:D200,"Partial")', PARTIAL_FILL, '9C6500'),
    ('Not Built', '=COUNTIF(D5:D200,"Not Built")', NOT_FILL, '9C0006'),
    ('Total Items', '=COUNTIF(D5:D200,"Done")+COUNTIF(D5:D200,"Partial")+COUNTIF(D5:D200,"Not Built")', PatternFill('solid', fgColor='E0E0E0'), '333333'),
    ('Completion %', '=ROUND(COUNTIF(D5:D200,"Done")/(COUNTIF(D5:D200,"Done")+COUNTIF(D5:D200,"Partial")+COUNTIF(D5:D200,"Not Built"))*100,1)&"%"', PatternFill('solid', fgColor='E0E0E0'), '1B3A5C'),
]

for label, formula, fill, color in summary:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c1 = ws.cell(row=row, column=1, value=label)
    c1.font = Font(name='Arial', size=10, bold=True, color=color)
    c1.fill = fill
    c1.border = BORDER
    ws.cell(row=row, column=2).border = BORDER
    ws.cell(row=row, column=2).fill = fill
    c2 = ws.cell(row=row, column=3, value=formula)
    c2.font = Font(name='Arial', size=10, bold=True, color=color)
    c2.fill = fill
    c2.border = BORDER
    row += 1

# Phase breakdown
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
c = ws.cell(row=row, column=1, value='Phase-wise Breakdown')
c.font = Font(name='Arial', bold=True, size=12, color='1B3A5C')
row += 1

for hdr_col, hdr_val in enumerate(['Phase', 'Description', 'Items', 'Done', 'Status'], 1):
    c = ws.cell(row=row, column=hdr_col, value=hdr_val)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER
row += 1

phases = [
    ('1', 'Foundation & Auth', '=COUNTIF(E5:E200,"1")', '=COUNTIFS(E5:E200,"1",D5:D200,"Done")', ''),
    ('2A', 'Employee & Org Mgmt', '=COUNTIF(E5:E200,"2A")', '=COUNTIFS(E5:E200,"2A",D5:D200,"Done")', ''),
    ('2B', 'Weekly Cycle', '=COUNTIF(E5:E200,"2B")', '=COUNTIFS(E5:E200,"2B",D5:D200,"Done")', ''),
    ('2C', 'Monthly Scoring & Tasks', '=COUNTIF(E5:E200,"2C")', '=COUNTIFS(E5:E200,"2C",D5:D200,"Done")', ''),
    ('2D', 'Review Lifecycle & Notifications', '=COUNTIF(E5:E200,"2D")', '=COUNTIFS(E5:E200,"2D",D5:D200,"Done")', ''),
    ('3', 'Automation & Escalation', '=COUNTIF(E5:E200,"3")', '=COUNTIFS(E5:E200,"3",D5:D200,"Done")', ''),
    ('4', 'History, Feedback & Reports', '=COUNTIF(E5:E200,"4")', '=COUNTIFS(E5:E200,"4",D5:D200,"Done")', ''),
    ('5', 'Annual Increments', '=COUNTIF(E5:E200,"5")', '=COUNTIFS(E5:E200,"5",D5:D200,"Done")', ''),
    ('6', 'Executive & Edge Cases', '=COUNTIF(E5:E200,"6")', '=COUNTIFS(E5:E200,"6",D5:D200,"Done")', ''),
]

for phase, desc, items_f, done_f, status in phases:
    r = row
    ws.cell(row=r, column=1, value=phase).font = BOLD_FONT
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value=desc).font = BODY_FONT
    ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=3, value=items_f).font = BODY_FONT
    ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=3).border = BORDER
    ws.cell(row=r, column=4, value=done_f).font = BODY_FONT
    ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=4).border = BORDER
    status_f = f'=IF(C{r}=D{r},"Complete",IF(D{r}=0,"Not Started","In Progress"))'
    sc = ws.cell(row=r, column=5, value=status_f)
    sc.font = BOLD_FONT
    sc.alignment = Alignment(horizontal='center')
    sc.border = BORDER
    row += 1

ws.sheet_properties.pageSetUpPr = None
ws.freeze_panes = 'A5'

out = '/sessions/loving-adoring-clarke/mnt/PMS/Crystal_People_SOP_Tracker.xlsx'
wb.save(out)
print(f'Saved to {out}')
