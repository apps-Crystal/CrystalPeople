from openpyxl import load_workbook

wb = load_workbook('/sessions/loving-adoring-clarke/mnt/PMS/Crystal_People_SOP_Tracker.xlsx')
ws = wb['SOP Build Status']

# Find last data row (before summary)
last_data = 0
for r in range(5, 90):
    v = ws.cell(row=r, column=4).value
    if v in ('Done', 'Partial', 'Not Built'):
        last_data = r

print(f'Last data row: {last_data}')

phases = [
    (98, '1'), (99, '2A'), (100, '2B'), (101, '2C'), (102, '2D'),
    (103, '3'), (104, '4'), (105, '5'), (106, '6'),
]

rng = f'E5:E{last_data}'
drng = f'D5:D{last_data}'

for row, phase in phases:
    ws.cell(row=row, column=3).value = f'=COUNTIF({rng},"{phase}")'
    ws.cell(row=row, column=4).value = f'=COUNTIFS({rng},"{phase}",{drng},"Done")'
    ws.cell(row=row, column=5).value = f'=IF(C{row}=0,"N/A",IF(C{row}=D{row},"Complete",IF(D{row}=0,"Not Started","In Progress")))'

# Also fix summary formulas
for r in range(88, 95):
    v = ws.cell(row=r, column=3).value
    if isinstance(v, str) and 'COUNTIF' in v:
        ws.cell(row=r, column=3).value = v.replace('D5:D200', f'D5:D{last_data}').replace('E5:E200', f'E5:E{last_data}')

wb.save('/sessions/loving-adoring-clarke/mnt/PMS/Crystal_People_SOP_Tracker.xlsx')
print('Fixed ranges')
