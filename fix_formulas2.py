from openpyxl import load_workbook

wb = load_workbook('/sessions/loving-adoring-clarke/mnt/PMS/Crystal_People_SOP_Tracker.xlsx')
ws = wb['SOP Build Status']

# Debug: check what's in column E around phase rows
for r in range(96, 107):
    print(f'Row {r}: col3={ws.cell(row=r, column=3).value} col4={ws.cell(row=r, column=4).value} col5={ws.cell(row=r, column=5).value}')
