from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = load_workbook(''/sessions/loving-adoring-clarke/mnt/PMS/Crystal_People_SOP_Tracker_v2.xlsx')

ws = wb.create_sheet('Phase Roadmap')

HEADER_FILL = PatternFill('solid', fgColor='1B3A5C')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
thin = Side(style='thin', color='B0B0B0')
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)
DONE_FILL = PatternFill('solid', fgColor='C6EFCE')
DONE_FONT = Font(name='Arial', size=10, bold=True, color='006100')
IP_FILL = PatternFill('solid', fgColor='FFEB9C')
IP_FONT = Font(name='Arial', size=10, bold=True, color='9C6500')
NS_FILL = PatternFill('solid', fgColor='FFC7CE')
NS_FONT = Font(name='Arial', size=10, bold=True, color='9C0006')
PHASE_FILL = PatternFill('solid', fgColor='D6E4F0')
PHASE_FONT = Font(name='Arial', bold=True, size=11, color='1B3A5C')

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 55
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 40

ws.merge_cells('A1:F1')
ws['A1'] = 'Crystal People — Phase-wise Development Roadmap'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1B3A5C')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:F2')
ws['A2'] = 'SOP: HR-SOP-002 v1.0  |  As of: 26 March 2026'
ws['A2'].font = Font(name='Arial', size=9, italic=True, color='666666')
ws['A2'].alignment = Alignment(horizontal='center')

headers = ['Phase', 'Phase Name', 'Feature', 'Status', 'Go-Live Need', 'Notes']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
ws.row_dimensions[4].height = 24

phases = [
    ('Phase 1', 'Foundation & Auth', 'Complete', [
        ('Next.js 15 + TypeScript + Tailwind CSS setup', 'Done', 'Yes', ''),
        ('Google Sheets API integration (read/write/update/batch/cache)', 'Done', 'Yes', ''),
        ('Sheet schema validation system', 'Done', 'Yes', ''),
        ('JWT email/password authentication (login, logout, session)', 'Done', 'Yes', ''),
        ('4 roles: employee, manager, hr, md', 'Done', 'Yes', ''),
        ('First login password change flow', 'Done', 'Yes', ''),
        ('Forgot password with reset token', 'Done', 'Yes', ''),
    ]),
    ('Phase 2A', 'Employee & Org Management', 'In Progress', [
        ('Employee list + search (HR/MD)', 'Done', 'Yes', ''),
        ('Employee profile view + edit', 'Done', 'Yes', ''),
        ('Manager assignment (Manager_ID)', 'Done', 'Yes', ''),
        ('Org Changes audit log', 'Done', 'Yes', 'What changed, who, when'),
        ('Employee exit flow (mark inactive)', 'Partial', 'Nice to have', 'Status field exists, no dedicated UI'),
        ('Config management (current_month, year, windows)', 'Done', 'Yes', ''),
    ]),
    ('Phase 2B', 'Weekly Cycle', 'Complete', [
        ('Employee weekly reflection (accomplishments, plan, blockers, mood)', 'Done', 'Yes', 'SOP §6.1'),
        ('Manager weekly check-in per employee (5 fields)', 'Done', 'Yes', 'SOP §6.2'),
        ('Concern field visible to HR/MD only', 'Done', 'Yes', 'SOP §6.2 + §11'),
        ('Reflection acknowledgement by manager', 'Done', 'Yes', 'SOP §6.3'),
        ('Dashboard with weekly status cards', 'Done', 'Yes', ''),
        ('Pending actions widget', 'Done', 'Yes', ''),
        ('Role-based sidebar navigation', 'Done', 'Yes', ''),
    ]),
    ('Phase 2C', 'Monthly Scoring & Tasks', 'Complete', [
        ('Employee self-scoring (goals + 3 behaviours for white-collar)', 'Done', 'Yes', 'SOP §7.1'),
        ('Employee self-scoring (tasks + 3 behaviours for blue-collar)', 'Done', 'Yes', 'SOP §7.1'),
        ('Goal management (create, edit, drop with reason)', 'Done', 'Yes', 'SOP §4.1'),
        ('Min 3 / max 5 goals enforcement', 'Done', 'Yes', 'SOP §4.1'),
        ('Task management API (CRUD, activate/deactivate)', 'Done', 'Yes', 'SOP §4.2'),
        ('Score dimensions engine (getScoreDimensions)', 'Done', 'Yes', ''),
        ('computeAverage + getRatingBand + getRatingColor', 'Done', 'Yes', 'SOP §5'),
        ('Window enforcement (isWindowOpen per role)', 'Done', 'Yes', 'SOP §7 Day timeline'),
        ('Manager scoring with review notes + next month focus', 'Done', 'Yes', 'SOP §7.2'),
        ('Locked Average + Flagged Dimensions calculation', 'Done', 'Yes', 'SOP §7.1'),
        ('≥2 point gap flagged for discussion', 'Done', 'Yes', 'SOP §7.1'),
        ('Score My Team page (team list + individual scoring)', 'Done', 'Yes', ''),
        ('Team Tasks UI page', 'Done', 'Yes', 'Manager assigns tasks to blue-collar'),
        ('Team Goals page', 'Done', 'Yes', ''),
        ('Self-Score page', 'Done', 'Yes', ''),
        ('My Goals page', 'Done', 'Yes', ''),
        ('My Tasks page (employee view)', 'Done', 'Yes', ''),
        ('Team Overview page', 'Done', 'Yes', ''),
    ]),
    ('Phase 2D', 'Review Lifecycle & Notifications', 'In Progress', [
        ('Employee acknowledgement API', 'Done', 'Yes', 'SOP §7 "After manager submits"'),
        ('My Review page (score comparison, comments, acknowledge)', 'Done', 'Yes', ''),
        ('AI review notes validation (harsh/vague/unconstructive)', 'Done', 'Yes', 'SOP §7.2'),
        ('Weekly context in monthly scoring screen', 'Done', 'Yes', 'SOP §7.3'),
        ('HR Review Monitor page (all employees, status, lock)', 'Done', 'Yes', ''),
        ('Lock Cycle API (single + bulk lock)', 'Done', 'Yes', 'SOP §7 "Day 11"'),
        ('Notifications foundation (create, list, read, mark read)', 'Done', 'Yes', 'SOP §13'),
        ('NotificationBell component in header', 'Done', 'Yes', ''),
        ('Notification triggers (manager-score, acknowledge, lock)', 'Done', 'Yes', ''),
        ('HR Notifications admin page', 'Not Built', 'Nice to have', 'HR can use Review Monitor'),
    ]),
    ('Phase 3', 'Automation & Escalation', 'Not Started', [
        ('Day 1 — auto notify window open', 'Not Built', 'Nice to have', 'HR can message manually'),
        ('Day 3 — first auto reminder to non-scorers', 'Not Built', 'Nice to have', 'HR sends from Monitor page'),
        ('Day 6 — second auto reminder', 'Not Built', 'Nice to have', ''),
        ('Day 7 — auto-flag pending to HR', 'Not Built', 'No', 'Monitor page already shows all'),
        ('Day 10 — pending visible to MD + manager penalty', 'Not Built', 'No', 'Manual process for month 1'),
        ('Day 11 — auto window close + auto lock', 'Not Built', 'Nice to have', 'HR can bulk lock manually'),
        ('Saturday auto reminder for missed reflections', 'Not Built', 'Nice to have', 'HR sends manually'),
        ('Monday — missed reflections visible to HR', 'Not Built', 'No', 'Can check sheet directly'),
        ('Missed reflections tracked, visible to MD', 'Not Built', 'No', ''),
        ('Manager deadline escalation to dept head', 'Not Built', 'No', ''),
        ('Full escalation dashboard', 'Not Built', 'No', ''),
    ]),
    ('Phase 4', 'History, Feedback & Reports', 'Not Started', [
        ('My History page (12-month score trend)', 'Not Built', 'Nice to have', 'Only 1 month of data at launch'),
        ('My Feedback page (all past manager feedback)', 'Not Built', 'Nice to have', 'My Review covers current month'),
        ('HR Reports (dept-wise completion, score distribution)', 'Not Built', 'Nice to have', 'Can export from Google Sheets'),
        ('Excel bulk upload for employee data', 'Not Built', 'Nice to have', 'Manual entry or sheet edit'),
    ]),
    ('Phase 5', 'Annual Increments', 'Not Started', [
        ('12-month average score calculation', 'Not Built', 'No', 'Not needed until March 2027'),
        ('MD sets increment percentage bands', 'Not Built', 'No', ''),
        ('HR computes individual increments', 'Not Built', 'No', ''),
        ('MD approves full increment sheet', 'Not Built', 'No', ''),
        ('Increment letter generation', 'Not Built', 'No', ''),
    ]),
    ('Phase 6', 'Executive & Edge Cases', 'Not Started', [
        ('Executive Dashboard (org-wide summary for MD)', 'Not Built', 'No', 'MD uses HR Admin pages'),
        ('Rating Overrides by MD', 'Not Built', 'No', 'Can edit sheet directly'),
        ('Approve Increments page', 'Not Built', 'No', 'Phase 5 dependency'),
        ('Employee disputes rating — HR mediation flow', 'Not Built', 'No', 'Manual process'),
        ('Goal/task change request flow', 'Not Built', 'No', 'Manual process'),
        ('Employee refuses to acknowledge — HR documents', 'Not Built', 'No', 'Manual process'),
        ('Org edge cases (joiner cutoff, manager exit, reporting change)', 'Not Built', 'No', 'Manual handling'),
    ]),
]

row = 5
for phase_name, phase_desc, phase_status, features in phases:
    # Phase header row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c1 = ws.cell(row=row, column=1, value=f'{phase_name} — {phase_desc}')
    c1.font = PHASE_FONT
    c1.fill = PHASE_FILL
    c1.alignment = Alignment(vertical='center')
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = PHASE_FILL
        ws.cell(row=row, column=col).border = BORDER
    sc = ws.cell(row=row, column=4, value=phase_status)
    sc.font = Font(name='Arial', bold=True, size=10)
    sc.alignment = Alignment(horizontal='center', vertical='center')
    if phase_status == 'Complete':
        sc.fill = DONE_FILL
        sc.font = DONE_FONT
    elif phase_status == 'In Progress':
        sc.fill = IP_FILL
        sc.font = IP_FONT
    elif phase_status == 'Not Started':
        sc.fill = NS_FILL
        sc.font = NS_FONT
    ws.row_dimensions[row].height = 24
    row += 1

    for feat, status, golive, notes in features:
        ws.cell(row=row, column=1, value='').border = BORDER
        ws.cell(row=row, column=2, value='').border = BORDER
        c3 = ws.cell(row=row, column=3, value=feat)
        c3.font = BODY_FONT
        c3.border = BORDER
        c3.alignment = Alignment(vertical='center', wrap_text=True)

        c4 = ws.cell(row=row, column=4, value=status)
        c4.alignment = Alignment(horizontal='center', vertical='center')
        c4.border = BORDER
        if status == 'Done':
            c4.fill = DONE_FILL
            c4.font = DONE_FONT
        elif status == 'Partial':
            c4.fill = IP_FILL
            c4.font = IP_FONT
        elif status == 'Not Built':
            c4.fill = NS_FILL
            c4.font = NS_FONT

        c5 = ws.cell(row=row, column=5, value=golive)
        c5.font = BODY_FONT
        c5.border = BORDER
        c5.alignment = Alignment(horizontal='center', vertical='center')
        if golive == 'Yes':
            c5.fill = DONE_FILL
            c5.font = Font(name='Arial', size=10, color='006100')
        elif golive == 'Nice to have':
            c5.fill = IP_FILL
            c5.font = Font(name='Arial', size=10, color='9C6500')
        elif golive == 'No':
            c5.fill = PatternFill('solid', fgColor='E0E0E0')
            c5.font = Font(name='Arial', size=10, color='666666')

        c6 = ws.cell(row=row, column=6, value=notes)
        c6.font = Font(name='Arial', size=9, color='666666')
        c6.border = BORDER
        c6.alignment = Alignment(vertical='center', wrap_text=True)
        row += 1

# Legend
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws.cell(row=row, column=1, value='Go-Live Need Legend').font = Font(name='Arial', bold=True, size=11, color='1B3A5C')
row += 1
legend = [
    ('Yes', 'Must have — already built and working', DONE_FILL, '006100'),
    ('Nice to have', 'Useful but HR can work around it manually for month 1', IP_FILL, '9C6500'),
    ('No', 'Not needed for April 1 launch — can be built later', PatternFill('solid', fgColor='E0E0E0'), '666666'),
]
for label, desc, fill, color in legend:
    c1 = ws.cell(row=row, column=1, value=label)
    c1.font = Font(name='Arial', size=10, bold=True, color=color)
    c1.fill = fill
    c1.border = BORDER
    c1.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c2 = ws.cell(row=row, column=2, value=desc)
    c2.font = Font(name='Arial', size=10, color='333333')
    c2.border = BORDER
    for col in range(2, 7):
        ws.cell(row=row, column=col).border = BORDER
    row += 1

ws.freeze_panes = 'A5'

wb.save(''/sessions/loving-adoring-clarke/mnt/PMS/Crystal_People_SOP_Tracker_v2.xlsx')
print('Phase Roadmap tab added')
