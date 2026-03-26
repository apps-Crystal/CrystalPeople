from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

wb = load_workbook('/sessions/loving-adoring-clarke/mnt/PMS/Crystal_People_SOP_Tracker.xlsx')
ws_main = wb['SOP Build Status']

HEADER_FILL = PatternFill('solid', fgColor='1B3A5C')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
DONE_FILL = PatternFill('solid', fgColor='C6EFCE')
NOT_FILL = PatternFill('solid', fgColor='FFC7CE')
PARTIAL_FILL = PatternFill('solid', fgColor='FFEB9C')
BODY_FONT = Font(name='Arial', size=10)
thin = Side(style='thin', color='B0B0B0')
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

# Collect data rows from main sheet
done_rows = []
pending_rows = []

for row in range(5, ws_main.max_row + 1):
    status_cell = ws_main.cell(row=row, column=4).value
    if status_cell in ('Done', 'Partial', 'Not Built'):
        row_data = []
        for col in range(1, 7):
            row_data.append(ws_main.cell(row=row, column=col).value)
        if status_cell == 'Done':
            done_rows.append(row_data)
        else:
            pending_rows.append(row_data)

def create_tab(wb, name, rows, title_text, title_color):
    ws = wb.create_sheet(name)
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 52
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 45

    ws.merge_cells('A1:F1')
    ws['A1'] = title_text
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color=title_color)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Total: {len(rows)} items'
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['S.No', 'SOP Section', 'Feature / Requirement', 'Status', 'Phase', 'Remarks']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER

    for i, rd in enumerate(rows):
        r = i + 5
        for col, v in enumerate(rd, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if col == 1:
                c.alignment = Alignment(horizontal='center', vertical='center')
            if col == 4:
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.font = Font(name='Arial', size=10, bold=True)
                if v == 'Done':
                    c.fill = DONE_FILL
                    c.font = Font(name='Arial', size=10, bold=True, color='006100')
                elif v == 'Partial':
                    c.fill = PARTIAL_FILL
                    c.font = Font(name='Arial', size=10, bold=True, color='9C6500')
                elif v == 'Not Built':
                    c.fill = NOT_FILL
                    c.font = Font(name='Arial', size=10, bold=True, color='9C0006')
            if col == 5:
                c.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'A5'
    return ws

create_tab(wb, 'Completed', done_rows, 'Crystal People — Completed Features', '006100')
create_tab(wb, 'Pending', pending_rows, 'Crystal People — Pending / Not Built', '9C0006')

out = '/sessions/loving-adoring-clarke/mnt/PMS/Crystal_People_SOP_Tracker.xlsx'
wb.save(out)
print(f'Saved with 3 tabs: {wb.sheetnames}')
